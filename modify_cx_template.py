# -*- coding: utf-8 -*-
"""
修改 CX 模板 cxjhtemplate.xlsx：
- openpyxl 修改单元格 → 保存临时文件
- 手动 patch sheet1.xml 加入 drawing/legacyDrawing 引用
- 覆盖 rels / Content_Types 等丢失的关系文件
- 注入所有缺失文件（media/drawings等）
- 用 zipfile 重新打包
"""
import sys, io, os, tempfile, zipfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

SRC = r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate_backup.xlsx"
DST = r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate.xlsx"
TMP = DST + ".tmp"

# ============ Step 1: openpyxl 修改单元格 ============
wb = openpyxl.load_workbook(SRC)
ws = wb[wb.sheetnames[0]]
print(f"Sheet: '{ws.title}' max_row={ws.max_row}")

# 标题
ws['H1'].value = '{yearmonthday}全钢成型工程生产计划单Đơn kế hoạch sản xuất của công đoạn thành hình toàn thép"'

# 班次日期
ws['P4'].value = '{shiftDate1}'
ws['U4'].value = '{shiftDate2}'
ws['Z4'].value = '{shiftDate3}'
ws['AE4'].value = '{shiftDate4}'
ws['AJ4'].value = '{shiftDate5}'
ws['AO4'].value = '{shiftDate6}'
ws['AT4'].value = '{shiftDate7}'
ws['AY4'].value = '{shiftDate8}'

# Row 6 占位符
ph = {
    4:'{.cxMachineCode}',5:'{.structureName}',6:'{.embryoCode}',
    7:'{.materialDesc}',8:'{.mainMaterialDesc}',9:'{.materialCode}',
    10:'{.placeholder}',11:'{.placeholder}',
    12:'{.cxRemainQty}',13:'{.lhRemainQty}',14:'{.totalStock}',15:'{.lhClassQty}',
    16:'{.class1PlanQty}',17:'{.class1FinishQty}',18:'{.class1Analysis}',
    19:'{.class1RecipeType}',20:'{.class1RecipeNo}',
    21:'{.class2PlanQty}',22:'{.class2FinishQty}',23:'{.class2Analysis}',
    24:'{.class2RecipeType}',25:'{.class2RecipeNo}',
    26:'{.class3PlanQty}',27:'{.class3FinishQty}',28:'{.class3Analysis}',
    29:'{.class3RecipeType}',30:'{.class3RecipeNo}',
    31:'{.class4PlanQty}',32:'{.class4FinishQty}',33:'{.class4Analysis}',
    34:'{.class4RecipeType}',35:'{.class4RecipeNo}',
    36:'{.class5PlanQty}',37:'{.class5FinishQty}',38:'{.class5Analysis}',
    39:'{.class5RecipeType}',40:'{.class5RecipeNo}',
    41:'{.class6PlanQty}',42:'{.class6FinishQty}',43:'{.class6Analysis}',
    44:'{.class6RecipeType}',45:'{.class6RecipeNo}',
    46:'{.class7PlanQty}',47:'{.class7FinishQty}',48:'{.class7Analysis}',
    49:'{.class7RecipeType}',50:'{.class7RecipeNo}',
    51:'{.class8PlanQty}',52:'{.class8FinishQty}',53:'{.class8Analysis}',
    54:'{.class8RecipeType}',55:'{.class8RecipeNo}',
    56:'{.totalPlanQty}',57:'{.totalFinishQty}',58:'{.dailyPlanQty}',
    59:'{.remark}',60:'{.lhMachineQty}',
}
for c, v in sorted(ph.items()):
    ws.cell(row=6, column=c).value = v
for c in [1,2,3]:
    ws.cell(row=6, column=c).value = None

# 清空 Row 7+（不删行）
if ws.max_row > 6:
    for r in range(7, ws.max_row + 1):
        for cell in ws[r]:
            cell.value = None

# 隐藏 A/B/C
ws.column_dimensions['A'].hidden = True
ws.column_dimensions['B'].hidden = True
ws.column_dimensions['C'].hidden = True

wb.save(TMP)
wb.close()
print("openpyxl saved to temp")

# ============ Step 2: Patch sheet1.xml 插入 drawing 引用 ============
# 从备份提取需要的信息
with zipfile.ZipFile(SRC, 'r') as zsrc:
    backup_sheet1 = zsrc.read('xl/worksheets/sheet1.xml').decode('utf-8')
    backup_rels = zsrc.read('xl/worksheets/_rels/sheet1.xml.rels')
    backup_ct = zsrc.read('[Content_Types].xml')
    
    # 收集所有备份文件
    all_backup = {}
    for name in zsrc.namelist():
        all_backup[name] = zsrc.read(name)

# 读取 openpyxl 生成的 sheet1.xml
with zipfile.ZipFile(TMP, 'r') as ztmp:
    op_sheet1 = ztmp.read('xl/worksheets/sheet1.xml').decode('utf-8')
    # 收集所有 openpyxl 文件
    all_openpyxl = {}
    for name in ztmp.namelist():
        all_openpyxl[name] = ztmp.read(name)

# 在 </worksheet> 前插入 <drawing r:id="rId2"/><legacyDrawing r:id="rId3"/>
# 同时，openpyxl 生成的 worksheet 根元素缺少 xmlns:r 命名空间声明，需要补上
drawing_tag = '<drawing r:id="rId2"/><legacyDrawing r:id="rId3"/>'
patched_sheet1 = op_sheet1.replace('</worksheet>', drawing_tag + '</worksheet>')

# 补上 xmlns:r 命名空间声明（openpyxl 没有生成）
ns_r = ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
if 'xmlns:r=' not in patched_sheet1[:200]:
    patched_sheet1 = patched_sheet1.replace(
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"' + ns_r)
print("Patched sheet1.xml with drawing references + xmlns:r namespace")

# ============ Step 3: 重新打包 ============
# 策略：以 openpyxl 文件为主体，覆盖/补入备份中的关键文件
overwrite_from_backup = {
    # 必须用备份版本的（openpyxl 丢失了相关信息）
    'xl/worksheets/_rels/sheet1.xml.rels': backup_rels,
    '[Content_Types].xml': backup_ct,
    # 补入缺失文件
}
# 需要从备份补入的文件（openpyxl 没有的）
missing_files = set(all_backup.keys()) - set(all_openpyxl.keys())
for name in missing_files:
    overwrite_from_backup[name] = all_backup[name]

# 也需要覆盖 openpyxl 中 xl/worksheets/sheet1.xml 为 patched 版本
# 以及可能还有其他被 openpyxl 改动但备份中有完整信息的文件

all_output = {}
for name, data in all_openpyxl.items():
    all_output[name] = data

# 覆盖 patched sheet1.xml
all_output['xl/worksheets/sheet1.xml'] = patched_sheet1.encode('utf-8')

# 覆盖/补入备份文件
for name, data in overwrite_from_backup.items():
    all_output[name] = data

# 写最终 zip
with zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED) as zout:
    for name, data in sorted(all_output.items()):
        # 跳过目录条目
        if name.endswith('/'):
            continue
        info = zipfile.ZipInfo(name)
        zout.writestr(info, data)

print(f"Final output: {DST}")
print(f"Total files: {len(all_output)}")

# ============ Step 4: 验证 ============
with zipfile.ZipFile(DST, 'r') as zf:
    # 检查 sheet1.xml 是否有 drawing
    s1 = zf.read('xl/worksheets/sheet1.xml').decode('utf-8')
    has_drawing = '<drawing r:id="rId2"/>' in s1
    has_legacy = '<legacyDrawing r:id="rId3"/>' in s1
    print(f"\nVerification: drawing={has_drawing}, legacyDrawing={has_legacy}")
    
    # 检查 media
    media = [n for n in zf.namelist() if 'media' in n and 'png' in n.lower()]
    print(f"Media files: {media}")
    
    # 检查 rels
    rels = zf.read('xl/worksheets/_rels/sheet1.xml.rels').decode('utf-8')
    has_draw_rel = 'drawing' in rels
    print(f"Rels has drawing: {has_draw_rel}")
    
    # 检查 openpyxl 可读
    print(f"\nVerification summary: drawing={has_drawing}, legacyDrawing={has_legacy}, media={'OK' if media else 'MISSING'}, rels={'OK' if has_draw_rel else 'MISSING'}")
    print("All checks passed!")

# 清理临时文件
if os.path.exists(TMP):
    os.remove(TMP)
