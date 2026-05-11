# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

SRC = r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate_backup.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb[wb.sheetnames[0]]

# 搜索标题
for r in range(1, 6):
    for c in range(1, min(ws.max_column + 1, 65)):
        v = ws.cell(row=r, column=c).value
        if v and ('kế hoạch' in str(v).lower() or 'toàn thép' in str(v).lower()):
            print(f"Row{r} C{c}: {v!r}")
            for mc in ws.merged_cells.ranges:
                if ws.cell(row=r, column=c).coordinate in mc:
                    print(f"  merged: {mc}")

# 同时也打 Row 1-3 完整内容
print("\n=== Row 1 ===")
for c in range(1, 20):
    v = ws.cell(row=1, column=c).value
    if v: print(f"  C{c}: {v!r}")

print("\n=== Row 2 ===")
for c in range(1, 20):
    v = ws.cell(row=2, column=c).value
    if v: print(f"  C{c}: {v!r}")

print("\n=== Row 3 ===")
for c in range(1, 20):
    v = ws.cell(row=3, column=c).value
    if v: print(f"  C{c}: {v!r}")

wb.close()
