# -*- coding: utf-8 -*-
"""精确分析 CX 模板每列的含义 —— Row 4 中文表头、Row 5 越南语表头、Row 6 示例数据值"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

SRC = r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate_backup.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb[wb.sheetnames[0]]

print("=" * 100)
print("列索引 | 中文表头(Row4)                          | 越南语(Row5)                             | 示例数据(Row6)")
print("=" * 100)

for col in range(1, ws.max_column + 1):
    h4 = ws.cell(row=4, column=col).value
    h5 = ws.cell(row=5, column=col).value
    d6 = ws.cell(row=6, column=col).value
    
    h4s = repr(h4).replace('\n', '\\n') if h4 else '(空)'
    h5s = repr(h5).replace('\n', '\\n') if h5 else '(空)'
    d6s = repr(d6)[:80] if d6 else '(空)'
    
    print(f"  C{col:<3} | {h4s:<43} | {h5s:<43} | {d6s}")

wb.close()
