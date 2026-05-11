# -*- coding: utf-8 -*-
"""详细分析 CX 模板行结构"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

path = r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate.xlsx"
wb = openpyxl.load_workbook(path)

print("All sheets:", wb.sheetnames)

# 分析第一个 sheet
ws = wb[wb.sheetnames[0]]
print(f"\nSheet 0: '{wb.sheetnames[0]}' max_row={ws.max_row}, max_col={ws.max_column}")

# 打印每一行的列号对应值映射
print("\n--- 逐行列映射 ---")
for row_idx in range(1, min(ws.max_row + 1, 20)):
    row = ws[row_idx]
    cols = {}
    for cell in row:
        if cell.value is not None:
            cols[cell.column] = repr(cell.value)
    if cols:
        # 用简短格式
        items = [f"C{c}={v}" for c, v in sorted(cols.items())]
        print(f"Row {row_idx}: " + " | ".join(items))
    else:
        print(f"Row {row_idx}: (empty)")

# 查找占位符
print("\n--- 占位符 ---")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        val = str(cell.value) if cell.value is not None else ""
        if '{' in val:
            print(f"  {cell.coordinate}: {cell.value}")

# 分析合并单元格
print("\n--- 合并单元格 ---")
for mc in sorted(ws.merged_cells.ranges, key=str):
    print(f"  {mc}")

wb.close()
