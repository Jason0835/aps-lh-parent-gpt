# -*- coding: utf-8 -*-
"""分析 LH 和 CX 模板第一页结构，找占位符"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

def analyze_sheet0(filepath, label):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    print(f"\n{'='*80}")
    print(f"Template: {label}")
    print(f"Sheet 0: '{wb.sheetnames[0]}'  rows={ws.max_row}, cols={ws.max_column}")
    
    # 打印所有行（前15行）
    print("\nAll rows (first 15):")
    for row_idx in range(1, min(ws.max_row + 1, 16)):
        row = ws[row_idx]
        cells = []
        for cell in row:
            val = cell.value
            if val is not None:
                cells.append(f"C{cell.column}={val!r}")
        print(f"  Row {row_idx}: {', '.join(cells) if cells else '(empty)'}")
    
    # 合并单元格
    print(f"\nMerged cells ({len(ws.merged_cells.ranges)}):")
    for mc in sorted(ws.merged_cells.ranges, key=str):
        print(f"  {mc}")
    
    # 查找占位符
    print("\nPlaceholders:")
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""
            if '{' in val:
                print(f"  Cell({cell.coordinate}): {val}")
                count += 1
    if count == 0:
        print("  (none)")
    
    wb.close()

analyze_sheet0(
    r"d:\Company\GIT\jy_aps\APS-Modules\aps-lh\src\main\resources\excelModel\lhjhtemplate.xlsx",
    "LH"
)

analyze_sheet0(
    r"d:\Company\GIT\jy_aps\APS-Modules\aps-cx\src\main\resources\excelModel\cxjhtemplate.xlsx",
    "CX"
)
